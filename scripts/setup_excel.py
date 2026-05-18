#!/usr/bin/env python3
"""Create a blank job tracker Excel workbook with the 29-column schema and
dropdown validation. Run once after cloning the repo:

    python scripts/setup_excel.py

Writes to $BASE_DIR/job_tracker.xlsx (path resolved from .env)."""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    ("Job ID",                                10),
    ("Market",                                14),
    ("Job Title",                             34),
    ("Company",                               22),
    ("Company Size",                          14),
    ("Location",                              22),
    ("Work Mode",                             12),
    ("Salary Min",                            12),
    ("Salary Max",                            12),
    ("Currency",                              10),
    ("Meets HQP Threshold (€40,077+)",        18),
    ("Meets DNV Threshold ($2,600/mo)",       18),
    ("Primary Visa Path",                     14),
    ("Required Tools",                        28),
    ("Spanish Required",                      14),
    ("Experience Required (years)",           18),
    ("Education Required",                    22),
    ("Key Responsibilities",                  40),
    ("Job Board",                             14),
    ("Job URL",                               40),
    ("Date Posted",                           12),
    ("Application Deadline",                  16),
    ("Date Applied",                          12),
    ("Application Status",                    16),
    ("Recruiter / Contact",                   24),
    ("Notes",                                 30),
    ("Match Score",                           12),
    ("Suggested Yearly",                      14),
    ("Suggested Monthly",                     14),
]

DROPDOWNS = {
    "Market":             '"Spain,US Remote,LATAM Agency"',
    "Work Mode":          '"On-site,Hybrid,Remote"',
    "Application Status": '"Not Applied,Applied,In Discussion,Interview,Offer,Rejected,Withdrawn"',
    "Primary Visa Path":  '"HQP,DNV,Either,N/A"',
}


def main():
    base = os.environ.get("BASE_DIR") or os.path.expanduser("~/job-search")
    out = os.path.join(base, "job_tracker.xlsx")
    os.makedirs(base, exist_ok=True)
    if os.path.exists(out):
        print(f"Refusing to overwrite existing {out}", file=sys.stderr)
        print("Delete it first if you really want a blank tracker.", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    # Title row
    ws.cell(1, 1, "Job Search Tracker").font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for i, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(2, i, label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header
    ws.freeze_panes = "A3"

    # Dropdowns
    for label, formula in DROPDOWNS.items():
        col_idx = next(i for i, (n, _) in enumerate(COLUMNS, start=1) if n == label)
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.add(f"{col_letter}3:{col_letter}1000")
        ws.add_data_validation(dv)

    # Autofilter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"

    wb.save(out)
    print(f"Created {out}")


if __name__ == "__main__":
    main()
