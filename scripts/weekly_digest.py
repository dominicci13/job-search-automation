#!/usr/bin/env python3
"""Weekly job search digest — runs every Friday at 19:00 via launchd.

Reads the Excel tracker, computes this-week metrics, generates a polished
matplotlib dashboard PNG, and sends an HTML email via Mail.app with the
dashboard embedded inline as a data-URI.
"""
import base64
import datetime
import os
import subprocess
import sys
import tempfile
from collections import Counter
from html import escape

import matplotlib
from openpyxl import load_workbook

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import rcParams


def _require_env(name):
    val = os.environ.get(name)
    if not val:
        raise SystemExit(
            f"{name} not set — source your .env before running (e.g. via run_weekly_digest.sh)"
        )
    return val


BASE_DIR = _require_env("BASE_DIR")
XLSX = os.path.join(BASE_DIR, "job_tracker.xlsx")
PNG_PATH = "/tmp/jobsearch_weekly_dashboard.png"
EMAIL_TO = _require_env("EMAIL_TO")

# Visual palette — used by both matplotlib and HTML
PALETTE = {
    "primary":      "#2563EB",
    "primary_dark": "#1E40AF",
    "primary_soft": "#DBEAFE",
    "amber":        "#F59E0B",
    "amber_soft":   "#FEF3C7",
    "green":        "#10B981",
    "green_soft":   "#D1FAE5",
    "red":          "#EF4444",
    "red_soft":     "#FEE2E2",
    "gray":         "#6B7280",
    "gray_soft":    "#F3F4F6",
    "ink":          "#0F172A",
    "muted":        "#64748B",
    "bg":           "#F8FAFC",
    "card":         "#FFFFFF",
    "border":       "#E5E7EB",
}

STATUS_COLOR = {
    "Applied":       PALETTE["primary"],
    "In Discussion": PALETTE["amber"],
    "Interview":     PALETTE["green"],
    "Offer":         "#059669",
    "Rejected":      PALETTE["gray"],
    "Unknown":       PALETTE["gray"],
}

# ──────────────────────────────────────────────────────────────────────
# Data loading + bucketing
# ──────────────────────────────────────────────────────────────────────


def parse_date(d):
    if not d:
        return None
    if isinstance(d, datetime.datetime):
        return d.date()
    if isinstance(d, datetime.date):
        return d
    if isinstance(d, str):
        try:
            return datetime.datetime.strptime(d[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def bucket_status(s):
    s = (s or "").strip()
    if not s:
        return "Unknown"
    sl = s.lower()
    if "not applied" in sl:
        return "Not Applied"
    if "rejected" in sl or "closed" in sl or "withdrawn" in sl:
        return "Rejected"
    if "offer" in sl:
        return "Offer"
    if "interview" in sl:
        return "Interview"
    if "discussion" in sl or "screening" in sl or "recruiter" in sl:
        return "In Discussion"
    if sl.startswith("applied"):
        return "Applied"
    return s


def bucket_visa(v):
    v = (v or "").strip()
    if "HQP" in v and ("DNV" in v or "Digital Nomad" in v):
        return "Either"
    if "HQP" in v:
        return "HQP"
    if "DNV" in v or "Digital Nomad" in v:
        return "DNV"
    if "Either" in v:
        return "Either"
    return "Other/Unknown"


def market_emoji(m):
    return {"Spain": "🇪🇸", "US Remote": "🇺🇸", "LATAM Agency": "🌎"}.get(m, "📍")


def load_rows():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Job Listings"]
    rows = []
    for r in range(3, ws.max_row + 1):
        if not ws.cell(row=r, column=1).value:
            continue
        rows.append({
            "id":           ws.cell(row=r, column=1).value,
            "market":       ws.cell(row=r, column=2).value or "",
            "title":        ws.cell(row=r, column=3).value or "",
            "company":      ws.cell(row=r, column=4).value or "",
            "location":     ws.cell(row=r, column=6).value or "",
            "sal_min":      ws.cell(row=r, column=8).value,
            "sal_max":      ws.cell(row=r, column=9).value,
            "currency":     ws.cell(row=r, column=10).value or "",
            "visa_path":    ws.cell(row=r, column=13).value or "",
            "board":        ws.cell(row=r, column=19).value or "",
            "url":          ws.cell(row=r, column=20).value or "",
            "date_posted":  parse_date(ws.cell(row=r, column=21).value),
            "deadline":     parse_date(ws.cell(row=r, column=22).value),
            "date_applied": parse_date(ws.cell(row=r, column=23).value),
            "status":       ws.cell(row=r, column=24).value or "",
            "recruiter":    ws.cell(row=r, column=25).value or "",
            "score":        ws.cell(row=r, column=27).value or 0,
        })
    return rows


# ──────────────────────────────────────────────────────────────────────
# Matplotlib dashboard
# ──────────────────────────────────────────────────────────────────────


def build_dashboard(apps_this, apps_last, weeks_trend, mon_this, fri_this):
    rcParams["font.family"] = ["-apple-system", "Helvetica Neue", "Helvetica", "Arial", "sans-serif"]
    rcParams["font.size"] = 10
    rcParams["axes.titleweight"] = "bold"
    rcParams["axes.titlesize"] = 11
    rcParams["axes.titlepad"] = 12
    rcParams["axes.spines.top"] = False
    rcParams["axes.spines.right"] = False
    rcParams["axes.edgecolor"] = PALETTE["border"]
    rcParams["axes.labelcolor"] = PALETTE["muted"]
    rcParams["xtick.color"] = PALETTE["muted"]
    rcParams["ytick.color"] = PALETTE["muted"]
    rcParams["grid.color"] = PALETTE["border"]
    rcParams["grid.alpha"] = 0.6

    fig, axs = plt.subplots(2, 2, figsize=(11.5, 8.0), facecolor="white")
    fig.suptitle(
        f"Weekly Job Search Dashboard  ·  {mon_this.strftime('%b %d')} – {fri_this.strftime('%b %d, %Y')}",
        fontsize=14, fontweight="bold", color=PALETTE["ink"], y=0.98,
    )

    # ── Panel 1: applications by market (this vs last) ──
    mc_this = Counter(r["market"] for r in apps_this)
    mc_last = Counter(r["market"] for r in apps_last)
    all_markets = sorted(set(list(mc_this.keys()) + list(mc_last.keys()))) or ["Spain", "US Remote", "LATAM Agency"]
    x_pos = list(range(len(all_markets)))
    width = 0.36
    axs[0, 0].bar([i - width/2 for i in x_pos], [mc_last.get(m, 0) for m in all_markets],
                  width, label="Last week", color="#CBD5E1", edgecolor="white", linewidth=0.8)
    axs[0, 0].bar([i + width/2 for i in x_pos], [mc_this.get(m, 0) for m in all_markets],
                  width, label="This week", color=PALETTE["primary"], edgecolor="white", linewidth=0.8)
    for i, m in enumerate(all_markets):
        if mc_this.get(m, 0) > 0:
            axs[0, 0].text(i + width/2, mc_this[m] + 0.05, str(mc_this[m]),
                           ha="center", va="bottom", fontsize=9, fontweight="bold", color=PALETTE["ink"])
        if mc_last.get(m, 0) > 0:
            axs[0, 0].text(i - width/2, mc_last[m] + 0.05, str(mc_last[m]),
                           ha="center", va="bottom", fontsize=9, color=PALETTE["muted"])
    axs[0, 0].set_xticks(x_pos)
    axs[0, 0].set_xticklabels(all_markets, fontsize=9)
    axs[0, 0].set_title("Applications by Market", loc="left")
    axs[0, 0].legend(fontsize=9, frameon=False, loc="upper right")
    axs[0, 0].grid(axis="y", alpha=0.3)
    axs[0, 0].set_axisbelow(True)

    # ── Panel 2: status donut ──
    status_counts = Counter(bucket_status(r["status"]) for r in apps_this)
    if status_counts:
        labels = list(status_counts.keys())
        sizes = list(status_counts.values())
        colors = [STATUS_COLOR.get(l, PALETTE["gray"]) for l in labels]
        wedges, _, autotexts = axs[0, 1].pie(
            sizes, labels=None, colors=colors, autopct="%1.0f%%",
            startangle=90, wedgeprops={"width": 0.45, "edgecolor": "white", "linewidth": 2},
            textprops={"fontsize": 10, "fontweight": "bold", "color": "white"},
        )
        axs[0, 1].legend(wedges, [f"{l} ({c})" for l, c in status_counts.items()],
                         loc="center left", bbox_to_anchor=(0.95, 0.5), frameon=False, fontsize=9)
        axs[0, 1].text(0, 0, f"{sum(sizes)}\napps", ha="center", va="center",
                       fontsize=14, fontweight="bold", color=PALETTE["ink"])
    else:
        axs[0, 1].text(0.5, 0.5, "No applications\nthis week",
                       ha="center", va="center", transform=axs[0, 1].transAxes,
                       fontsize=11, color=PALETTE["muted"])
        axs[0, 1].axis("off")
    axs[0, 1].set_title("Status — This Week", loc="left")

    # ── Panel 3: 4-week trend ──
    labels = [w[0] for w in weeks_trend]
    counts = [w[1] for w in weeks_trend]
    axs[1, 0].plot(range(len(labels)), counts, marker="o", color=PALETTE["primary"],
                   linewidth=2.4, markersize=9, markerfacecolor="white",
                   markeredgewidth=2.4, markeredgecolor=PALETTE["primary"])
    axs[1, 0].fill_between(range(len(counts)), counts, alpha=0.12, color=PALETTE["primary"])
    axs[1, 0].set_xticks(range(len(labels)))
    axs[1, 0].set_xticklabels(labels, fontsize=9)
    axs[1, 0].set_title("Applications per Week — 4-Week View", loc="left")
    axs[1, 0].grid(alpha=0.3, axis="y")
    axs[1, 0].set_axisbelow(True)
    for i, c in enumerate(counts):
        axs[1, 0].annotate(str(c), (i, c), textcoords="offset points", xytext=(0, 12),
                           ha="center", fontsize=10, fontweight="bold", color=PALETTE["ink"])
    if counts:
        axs[1, 0].set_ylim(-0.5, max(counts) + 1.8)

    # ── Panel 4: match score histogram ──
    scores = [r["score"] for r in apps_this if isinstance(r["score"], (int, float)) and r["score"]]
    if scores:
        score_counts = Counter(scores)
        score_keys = sorted(score_counts.keys())
        bar_colors = []
        for s in score_keys:
            if s >= 4:
                bar_colors.append(PALETTE["green"])
            elif s == 3:
                bar_colors.append(PALETTE["primary"])
            else:
                bar_colors.append("#94A3B8")
        bars = axs[1, 1].bar(score_keys, [score_counts[s] for s in score_keys],
                             color=bar_colors, edgecolor="white", linewidth=1.5, width=0.7)
        for b in bars:
            h = b.get_height()
            axs[1, 1].text(b.get_x() + b.get_width()/2, h + 0.05, str(int(h)),
                           ha="center", va="bottom", fontsize=10, fontweight="bold",
                           color=PALETTE["ink"])
        axs[1, 1].set_xticks([1, 2, 3, 4, 5])
        axs[1, 1].set_xlabel("Match Score (1–5)", fontsize=9)
        axs[1, 1].grid(axis="y", alpha=0.3)
        axs[1, 1].set_axisbelow(True)
        if score_counts:
            axs[1, 1].set_ylim(0, max(score_counts.values()) + 1)
    else:
        axs[1, 1].text(0.5, 0.5, "No scored apps\nthis week",
                       ha="center", va="center", transform=axs[1, 1].transAxes,
                       fontsize=11, color=PALETTE["muted"])
        axs[1, 1].axis("off")
    axs[1, 1].set_title("Match Score Distribution", loc="left")

    plt.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(PNG_PATH, dpi=120, bbox_inches="tight", facecolor="white")
    plt.close(fig)


# ──────────────────────────────────────────────────────────────────────
# HTML email body
# ──────────────────────────────────────────────────────────────────────


def _pct(c, total):
    return f"{(c / total * 100):.0f}%" if total else "0%"


def _badge(text, bg, fg):
    return (
        f'<span style="display:inline-block;padding:4px 12px;border-radius:14px;'
        f'background:{bg};color:{fg};font-size:12px;font-weight:600;'
        f'margin-right:6px;white-space:nowrap;">{escape(text)}</span>'
    )


def status_badge(label, count):
    color = STATUS_COLOR.get(label, PALETTE["gray"])
    soft_map = {
        "Applied":       PALETTE["primary_soft"],
        "In Discussion": PALETTE["amber_soft"],
        "Interview":     PALETTE["green_soft"],
        "Offer":         PALETTE["green_soft"],
        "Rejected":      PALETTE["gray_soft"],
        "Unknown":       PALETTE["gray_soft"],
    }
    soft = soft_map.get(label, PALETTE["gray_soft"])
    return (
        f'<span style="display:inline-block;padding:6px 14px;border-radius:16px;'
        f'background:{soft};color:{color};font-size:13px;font-weight:600;'
        f'margin:0 8px 8px 0;">{escape(label)} · {count}</span>'
    )


def score_pill(score):
    if score >= 4:
        bg, fg = PALETTE["green_soft"], PALETTE["green"]
    elif score == 3:
        bg, fg = PALETTE["primary_soft"], PALETTE["primary_dark"]
    else:
        bg, fg = PALETTE["gray_soft"], PALETTE["gray"]
    return (
        f'<span style="display:inline-block;padding:3px 10px;border-radius:10px;'
        f'background:{bg};color:{fg};font-size:12px;font-weight:700;">★ {score}</span>'
    )


def build_html_body(apps_this, apps_last, applied_all, weeks_trend, mon_this, fri_this, today):
    delta = len(apps_this) - len(apps_last)
    arrow = "↑" if delta > 0 else ("↓" if delta < 0 else "→")
    delta_text = f"+{delta}" if delta > 0 else (str(delta) if delta < 0 else "no change")

    # KPI tiles
    mc_this = Counter(r["market"] for r in apps_this)
    vc_this = Counter(bucket_visa(r["visa_path"]) for r in apps_this)
    scores = [r["score"] for r in apps_this if isinstance(r["score"], (int, float)) and r["score"]]
    avg_score = (sum(scores) / len(scores)) if scores else 0
    top_score = max(scores) if scores else 0

    # Pipeline
    pipeline_counts = Counter(bucket_status(r["status"]) for r in applied_all)
    active = sum(c for s, c in pipeline_counts.items() if s in {"Applied", "In Discussion", "Interview", "Offer"})

    # Top picks
    top_picks = sorted(apps_this, key=lambda r: -(r["score"] or 0))[:3]

    # Action items
    seven_d_ago = today - datetime.timedelta(days=7)
    thirty_d_ago = today - datetime.timedelta(days=30)
    fourteen_d_out = today + datetime.timedelta(days=14)
    stale_disc = [r for r in applied_all if bucket_status(r["status"]) == "In Discussion"
                  and r["date_applied"] and r["date_applied"] < seven_d_ago]
    stale_app = [r for r in applied_all if bucket_status(r["status"]) == "Applied"
                 and r["date_applied"] and r["date_applied"] < thirty_d_ago]
    deadlines = [r for r in applied_all if r["deadline"] and today <= r["deadline"] <= fourteen_d_out]

    # New recruiters
    new_recruiters = [r for r in apps_this if r["recruiter"]]

    # Embed PNG as base64
    with open(PNG_PATH, "rb") as f:
        png_b64 = base64.b64encode(f.read()).decode()

    # ──── Build HTML ────
    # Header card
    header = f"""
    <div style="background:linear-gradient(135deg,{PALETTE['primary_dark']} 0%,{PALETTE['primary']} 100%);
                color:white;padding:32px 36px;border-radius:14px 14px 0 0;">
      <div style="font-size:13px;opacity:0.85;letter-spacing:1px;text-transform:uppercase;font-weight:600;">
        Weekly Job Search Digest
      </div>
      <div style="font-size:14px;opacity:0.9;margin-top:4px;">
        {mon_this.strftime('%B %d')} – {fri_this.strftime('%B %d, %Y')}
      </div>
      <div style="font-size:56px;font-weight:700;margin-top:18px;line-height:1;">
        {len(apps_this)}
      </div>
      <div style="font-size:14px;opacity:0.9;margin-top:6px;">
        new application{'s' if len(apps_this) != 1 else ''} this week
        &nbsp;·&nbsp;
        <span style="color:white;font-weight:600;">{arrow} {abs(delta) if delta != 0 else 0} {delta_text} vs last week</span>
      </div>
    </div>
    """

    # KPI strip
    def kpi_tile(label, value, sub=""):
        return f"""
        <td style="width:33%;padding:0 6px;vertical-align:top;">
          <div style="background:{PALETTE['card']};border:1px solid {PALETTE['border']};
                      border-radius:12px;padding:18px 16px;">
            <div style="color:{PALETTE['muted']};font-size:11px;font-weight:600;
                        text-transform:uppercase;letter-spacing:0.6px;">{label}</div>
            <div style="color:{PALETTE['ink']};font-size:24px;font-weight:700;margin-top:6px;line-height:1.1;">{value}</div>
            <div style="color:{PALETTE['muted']};font-size:12px;margin-top:4px;">{sub}</div>
          </div>
        </td>
        """

    market_kpi = " · ".join(f"{market_emoji(m)} {c}" for m, c in mc_this.most_common()) or "—"
    visa_kpi = " · ".join(f"{v} {c}" for v, c in vc_this.most_common()) or "—"
    score_kpi = f"{avg_score:.1f}" if scores else "—"
    score_sub = f"top {top_score}/5" if scores else "no scores yet"

    kpi_strip = f"""
    <div style="padding:18px 12px 6px 12px;">
      <table style="width:100%;border-spacing:0;">
        <tr>
          {kpi_tile("By Market", market_kpi, "this week's split")}
          {kpi_tile("Visa Path", visa_kpi, "applications by route")}
          {kpi_tile("Match Score", score_kpi, score_sub)}
        </tr>
      </table>
    </div>
    """

    # Pipeline status
    status_html = "".join(
        status_badge(s, pipeline_counts[s])
        for s in ["Applied", "In Discussion", "Interview", "Offer", "Rejected"]
        if pipeline_counts.get(s, 0) > 0
    )
    pipeline_section = f"""
    <div style="padding:6px 24px 18px 24px;">
      <h3 style="color:{PALETTE['ink']};font-size:15px;font-weight:700;margin:18px 0 12px 0;
                 text-transform:uppercase;letter-spacing:0.6px;">
        Pipeline Status (all-time)
      </h3>
      <div>{status_html or '<span style="color:'+PALETTE['muted']+';font-size:13px;">No applications yet</span>'}</div>
      <div style="margin-top:10px;color:{PALETTE['muted']};font-size:13px;">
        Active pipeline (excluding Rejected): <strong style="color:{PALETTE['ink']};">{active}</strong>
      </div>
    </div>
    """

    # Top picks
    picks_cards = []
    for r in top_picks:
        market_tag = f'<span style="font-size:11px;color:{PALETTE["muted"]};">{market_emoji(r["market"])} {escape(r["market"])}</span>'
        status_color = STATUS_COLOR.get(bucket_status(r["status"]), PALETTE["gray"])
        title = escape(r["title"])[:80]
        company = escape(r["company"])[:60]
        picks_cards.append(f"""
        <div style="background:{PALETTE['card']};border:1px solid {PALETTE['border']};border-radius:10px;
                    padding:14px 16px;margin-bottom:10px;display:block;">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;">
            <div style="flex:1;min-width:0;">
              <div style="color:{PALETTE['ink']};font-size:14px;font-weight:600;line-height:1.4;">{title}</div>
              <div style="color:{PALETTE['muted']};font-size:13px;margin-top:4px;">@ {company}</div>
              <div style="margin-top:8px;">{market_tag}</div>
            </div>
            <div style="text-align:right;margin-left:12px;white-space:nowrap;">
              {score_pill(r["score"] or 0)}
              <div style="margin-top:6px;font-size:12px;color:{status_color};font-weight:600;">
                {escape(bucket_status(r["status"]))}
              </div>
            </div>
          </div>
        </div>
        """)
    if picks_cards:
        picks_section = f"""
        <div style="padding:6px 24px 18px 24px;">
          <h3 style="color:{PALETTE['ink']};font-size:15px;font-weight:700;margin:18px 0 12px 0;
                     text-transform:uppercase;letter-spacing:0.6px;">
            🏆 Top Picks This Week
          </h3>
          {''.join(picks_cards)}
        </div>
        """
    else:
        picks_section = ""

    # Action items
    action_blocks = []
    if stale_disc:
        items = "".join(
            f'<li style="margin-bottom:6px;color:{PALETTE["ink"]};font-size:13px;">'
            f'<strong>{escape(r["company"])}</strong> · applied {r["date_applied"]}'
            f'{" · " + escape(r["recruiter"]) if r["recruiter"] else ""}</li>'
            for r in stale_disc[:5]
        )
        action_blocks.append(f"""
        <div style="background:{PALETTE['amber_soft']};border-left:4px solid {PALETTE['amber']};
                    border-radius:8px;padding:14px 16px;margin-bottom:10px;">
          <div style="color:{PALETTE['ink']};font-weight:700;font-size:14px;margin-bottom:8px;">
            🔔 Follow-up · {len(stale_disc)} "In Discussion" job(s) idle 7+ days
          </div>
          <ul style="margin:0;padding-left:18px;">{items}</ul>
        </div>
        """)
    if stale_app:
        items = "".join(
            f'<li style="margin-bottom:6px;color:{PALETTE["ink"]};font-size:13px;">'
            f'<strong>{escape(r["company"])}</strong> · applied {r["date_applied"]}</li>'
            for r in stale_app[:5]
        )
        action_blocks.append(f"""
        <div style="background:{PALETTE['gray_soft']};border-left:4px solid {PALETTE['gray']};
                    border-radius:8px;padding:14px 16px;margin-bottom:10px;">
          <div style="color:{PALETTE['ink']};font-weight:700;font-size:14px;margin-bottom:8px;">
            💤 Stale · {len(stale_app)} "Applied" job(s) silent 30+ days
          </div>
          <ul style="margin:0;padding-left:18px;">{items}</ul>
          <div style="color:{PALETTE['muted']};font-size:12px;margin-top:8px;">
            These are likely silent rejections — consider marking as Rejected/Withdrawn.
          </div>
        </div>
        """)
    if deadlines:
        items = "".join(
            f'<li style="margin-bottom:6px;color:{PALETTE["ink"]};font-size:13px;">'
            f'<strong>{escape(r["company"])}</strong> · closes {r["deadline"]}</li>'
            for r in deadlines[:5]
        )
        action_blocks.append(f"""
        <div style="background:{PALETTE['red_soft']};border-left:4px solid {PALETTE['red']};
                    border-radius:8px;padding:14px 16px;margin-bottom:10px;">
          <div style="color:{PALETTE['ink']};font-weight:700;font-size:14px;margin-bottom:8px;">
            ⏰ Deadlines · {len(deadlines)} closing within 14 days
          </div>
          <ul style="margin:0;padding-left:18px;">{items}</ul>
        </div>
        """)
    if action_blocks:
        actions_section = f"""
        <div style="padding:6px 24px 18px 24px;">
          <h3 style="color:{PALETTE['ink']};font-size:15px;font-weight:700;margin:18px 0 12px 0;
                     text-transform:uppercase;letter-spacing:0.6px;">
            Action Items
          </h3>
          {''.join(action_blocks)}
        </div>
        """
    else:
        actions_section = f"""
        <div style="padding:6px 24px 6px 24px;">
          <div style="background:{PALETTE['green_soft']};border-left:4px solid {PALETTE['green']};
                      border-radius:8px;padding:14px 16px;color:{PALETTE['ink']};font-size:13px;">
            ✅ No urgent action items this week.
          </div>
        </div>
        """

    # Recruiters
    if new_recruiters:
        recruiter_cards = "".join(
            f'<div style="background:{PALETTE["card"]};border:1px solid {PALETTE["border"]};'
            f'border-radius:8px;padding:12px 14px;margin-bottom:8px;">'
            f'<div style="color:{PALETTE["ink"]};font-size:13px;font-weight:600;">{escape(r["recruiter"])[:120]}</div>'
            f'<div style="color:{PALETTE["muted"]};font-size:12px;margin-top:3px;">'
            f're: <strong>{escape(r["company"])}</strong></div></div>'
            for r in new_recruiters[:5]
        )
        recruiters_section = f"""
        <div style="padding:6px 24px 18px 24px;">
          <h3 style="color:{PALETTE['ink']};font-size:15px;font-weight:700;margin:18px 0 12px 0;
                     text-transform:uppercase;letter-spacing:0.6px;">
            📇 New Recruiter Contacts This Week
          </h3>
          {recruiter_cards}
        </div>
        """
    else:
        recruiters_section = ""

    # Trend bars (HTML)
    max_c = max((w[1] for w in weeks_trend), default=0) or 1
    trend_rows = "".join(
        f"""
        <tr>
          <td style="padding:6px 12px 6px 0;color:{PALETTE['muted']};font-size:12px;width:90px;white-space:nowrap;">{escape(label)}</td>
          <td style="padding:6px 0;">
            <div style="background:{PALETTE['border']};height:18px;border-radius:9px;width:100%;position:relative;overflow:hidden;">
              <div style="background:linear-gradient(90deg,{PALETTE['primary']},{PALETTE['primary_dark']});
                          height:100%;width:{(count/max_c)*100 if max_c else 0:.0f}%;border-radius:9px;"></div>
            </div>
          </td>
          <td style="padding:6px 0 6px 10px;color:{PALETTE['ink']};font-weight:700;font-size:13px;width:30px;text-align:right;">{count}</td>
        </tr>
        """
        for label, count in weeks_trend
    )
    trend_section = f"""
    <div style="padding:6px 24px 18px 24px;">
      <h3 style="color:{PALETTE['ink']};font-size:15px;font-weight:700;margin:18px 0 12px 0;
                 text-transform:uppercase;letter-spacing:0.6px;">
        📈 4-Week Trend
      </h3>
      <table style="width:100%;border-spacing:0;">{trend_rows}</table>
    </div>
    """

    # Dashboard PNG (embedded)
    dashboard_section = f"""
    <div style="padding:6px 24px 18px 24px;">
      <h3 style="color:{PALETTE['ink']};font-size:15px;font-weight:700;margin:18px 0 12px 0;
                 text-transform:uppercase;letter-spacing:0.6px;">
        📊 Visual Dashboard
      </h3>
      <img src="data:image/png;base64,{png_b64}"
           style="width:100%;max-width:920px;height:auto;border-radius:10px;
                  border:1px solid {PALETTE['border']};display:block;"
           alt="Weekly dashboard" />
    </div>
    """

    # Footer
    footer = f"""
    <div style="padding:18px 24px 24px 24px;border-top:1px solid {PALETTE['border']};color:{PALETTE['muted']};font-size:12px;line-height:1.6;">
      Tracker:&nbsp;<code style="background:{PALETTE['gray_soft']};padding:2px 6px;border-radius:4px;">{escape(XLSX)}</code><br>
      Generated automatically every Friday at 19:00. Reply to this email to flag any issues.
    </div>
    """

    full_html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:{PALETTE['bg']};
              font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,Arial,sans-serif;
              color:{PALETTE['ink']};">
  <div style="max-width:680px;margin:24px auto;background:{PALETTE['card']};
              border-radius:14px;box-shadow:0 2px 12px rgba(0,0,0,0.06);overflow:hidden;">
    {header}
    {kpi_strip}
    {pipeline_section}
    {picks_section}
    {actions_section}
    {recruiters_section}
    {trend_section}
    {dashboard_section}
    {footer}
  </div>
</body></html>"""
    return full_html


def build_plain_text_fallback(apps_this, apps_last, applied_all, weeks_trend, mon_this, fri_this):
    """Minimal plain-text fallback used by Mail clients that strip HTML."""
    delta = len(apps_this) - len(apps_last)
    lines = [
        f"WEEKLY JOB SEARCH DIGEST — {mon_this.strftime('%b %d')}-{fri_this.strftime('%b %d, %Y')}",
        "=" * 60,
        f"Applications this week: {len(apps_this)}  (last week: {len(apps_last)}, Δ {'+'+str(delta) if delta>0 else delta})",
        "",
        "View full report in the HTML version of this email.",
    ]
    return "\n".join(lines)


# ──────────────────────────────────────────────────────────────────────
# Email send via Mail.app
# ──────────────────────────────────────────────────────────────────────


def send_html_email(subject, html_body, plain_body):
    """Send via Mail.app, writing the AppleScript to a temp file to avoid the
    -e arg-size limit (HTML + base64 PNG can exceed several hundred KB)."""

    # AppleScript: read body from a tmp file (avoids escaping nightmares)
    body_tmp = tempfile.NamedTemporaryFile(
        delete=False, mode="w", suffix=".html", prefix="jobsearch_weekly_"
    )
    body_tmp.write(html_body)
    body_tmp.close()

    script_path = "/tmp/jobsearch_weekly_send.applescript"
    # CRITICAL: in Mail.app's AppleScript, when both `content` and `html content`
    # are set on a message, whichever is set LAST wins. Set html content last.
    applescript = f'''
set bodyFile to "{body_tmp.name}"
set htmlBody to do shell script "cat " & quoted form of bodyFile
tell application "Mail"
  set newMsg to make new outgoing message with properties ¬
    {{subject:"{subject.replace(chr(34), chr(92)+chr(34))}", visible:false}}
  tell newMsg
    make new to recipient at end of to recipients with properties {{address:"{EMAIL_TO}"}}
    -- Set plain-text content first so html content (set last) is the one used
    set content to "{plain_body.replace(chr(92), chr(92)+chr(92)).replace(chr(34), chr(92)+chr(34)).replace(chr(10), chr(92)+"n")}"
    set html content to htmlBody
  end tell
  send newMsg
end tell
'''
    with open(script_path, "w") as f:
        f.write(applescript)

    result = subprocess.run(["osascript", script_path], capture_output=True, text=True)
    os.unlink(body_tmp.name)
    if result.returncode != 0:
        raise RuntimeError(f"Email send failed: {result.stderr.strip()}")


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────


def main():
    today = datetime.date.today()
    weekday = today.weekday()
    mon_this = today - datetime.timedelta(days=weekday)
    fri_this = mon_this + datetime.timedelta(days=4)
    mon_last = mon_this - datetime.timedelta(days=7)
    fri_last = mon_last + datetime.timedelta(days=4)

    rows = load_rows()
    applied_all = [r for r in rows if r["date_applied"] is not None
                   and bucket_status(r["status"]) != "Not Applied"]
    apps_this = [r for r in applied_all if mon_this <= r["date_applied"] <= fri_this]
    apps_last = [r for r in applied_all if mon_last <= r["date_applied"] <= fri_last]

    weeks_trend = []
    for i in range(3, -1, -1):
        ws_start = mon_this - datetime.timedelta(days=7 * i)
        ws_end = ws_start + datetime.timedelta(days=4)
        label = f"{ws_start.strftime('%b %d')}–{ws_end.strftime('%d')}"
        cnt = sum(1 for r in applied_all if r["date_applied"] and ws_start <= r["date_applied"] <= ws_end)
        weeks_trend.append((label, cnt))

    build_dashboard(apps_this, apps_last, weeks_trend, mon_this, fri_this)
    html_body = build_html_body(apps_this, apps_last, applied_all, weeks_trend, mon_this, fri_this, today)
    plain_body = build_plain_text_fallback(apps_this, apps_last, applied_all, weeks_trend, mon_this, fri_this)

    print(f"This week: {len(apps_this)} applications, last week: {len(apps_last)}", file=sys.stderr)
    print(f"HTML body size: {len(html_body):,} chars (incl. base64 PNG)", file=sys.stderr)

    subject = f"Weekly Job Search Digest — {mon_this.strftime('%b %d')} ({len(apps_this)} new app{'s' if len(apps_this) != 1 else ''})"
    send_html_email(subject, html_body, plain_body)
    print(f"Digest emailed to {EMAIL_TO}", file=sys.stderr)


if __name__ == "__main__":
    main()
